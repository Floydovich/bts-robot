from unittest import TestCase

import openpyxl

from examples import EXAMPLE_COMPANY


class Reader:
    def __init__(self, path):
        self.path = path
        self.sheet = openpyxl.load_workbook(self.path).active

    def all(self):
        rows = []
        for row in self.sheet.iter_rows(min_row=6, min_col=1,
                                        max_row=self.sheet.max_row,
                                        max_col=self.sheet.max_column):
            rows.append(row)
        return rows


class ExcelReaderTest(TestCase):

    def setUp(self) -> None:
        path = "../list.xlsx"
        self.reader = Reader(path)

    def test_can_read_first_row(self):
        expected_row = EXAMPLE_COMPANY

        sheet = self.reader.sheet

        self.assertEqual(expected_row[0], list(sheet.rows)[5][1].value)

    def test_can_read_all_rows(self):
        rows = self.reader.all()

        self.assertEqual(712, len(rows))
        self.assertEqual(14, len(rows[0]))
        self.assertEqual(14, len(rows[-1]))
