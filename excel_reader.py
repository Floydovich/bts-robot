import openpyxl


class Reader:
    def __init__(self, path):
        self.path = path
        self.sheet = openpyxl.load_workbook(self.path).active

    def all_rows(self):
        rows = []
        for row in self.sheet.iter_rows(min_row=6, min_col=1,
                                        max_row=self.sheet.max_row,
                                        max_col=self.sheet.max_column):
            cells = [cell.value for cell in row]
            rows.append(cells)
        return rows
